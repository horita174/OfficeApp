from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

ws1 = wb.create_sheet("Mysheet",0)



# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

cell = ["A","B","C","D","E","F","G","H","I","J"]

for x in range(1,11):
        for y in range(1,11):
            ws1[cell[y - 1] + str(x)] = x * y

# Save the file
wb.save("sample.xlsx")
